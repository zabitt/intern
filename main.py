import openpyxl
from datetime import datetime, timedelta

def create_schedule(filename, start_date, num_parts, days_per_cycle):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Reading Schedule"

    # Write headers
    sheet['A1'] = "Date"
    sheet['B1'] = "Reading Part"
    sheet['C1'] = "Recall Part"

    current_date = start_date
    current_part = 1
    row = 2  # starting from row 2 to leave space for headers

    while current_part <= num_parts:
        # Writing schedule for current cycle
        for i in range(days_per_cycle):
            sheet.cell(row=row, column=1, value=current_date.strftime("%Y-%m-%d"))
            sheet.cell(row=row, column=2, value=f"Part {current_part}")
            sheet.cell(row=row, column=3, value=f"Recall Part {current_part - 1}" if current_part > 1 else "-")
            current_date += timedelta(days=1)
            row += 1
            current_part += 1
            if current_part > num_parts:
                break

        # Skipping a day for recall
        if current_part <= num_parts:
            sheet.cell(row=row, column=1, value=current_date.strftime("%Y-%m-%d"))
            sheet.cell(row=row, column=2, value="Recall")
            sheet.cell(row=row, column=3, value=f"Recall Part {current_part - 1}")
            current_date += timedelta(days=1)
            row += 1

    workbook.save(filename)
    print(f"Schedule saved to '{filename}'.")

# Example usage
start_date = datetime(2024, 1, 1)  # Start date for the schedule
num_parts = 60  # Total number of parts in the book
days_per_cycle = 3  # Number of days to read a part before recall

create_schedule("reading_schedule.xlsx", start_date, num_parts, days_per_cycle)
