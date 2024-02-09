from openpyxl import Workbook

# Create a new workbook
wb = Workbook()

# Select the active worksheet
ws = wb.active

# Add data to cells
ws["A1"] = 42
ws["B1"] = "Hello"
ws["C1"] = "World"

# Save the workbook
wb.save("example.xlsx")
